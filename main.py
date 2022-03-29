from helpers import format_header_to_db

import openpyxl
import os

wb = openpyxl.load_workbook(filename="pedidos.xlsx", data_only=True)
o_sheet = wb["sheet1"]

rows = []
header = []
check_header = True

for row_id in range(1, o_sheet.max_row):
	row = []
	for column_id in range(1, o_sheet.max_column):
		o_cell = o_sheet.cell(row=row_id, column=column_id)

		value = o_cell.value if o_cell.value != "None" else None
		if check_header:  # Fill the header
			header.append(value)
		else:
			row.append(value)

	if check_header:
		header = format_header_to_db(header)
		check_header = False

	if len(row) > 0:
		output = {}
		for i in row:
			index = row.index(i)

			header_name = header[index]
			output[header_name] = i

		rows.append(output)

print(rows)
